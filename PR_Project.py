import math
import openpyxl


# ____________KNN Algorithm____________


def KNN(dataSet, K, testInstance):

    global result, classes

    distList = []
    for i in range(len(dataSet)):

        distList.append(
            [
                "%.2f"
                % math.sqrt(
                    (testInstance[0] - dataSet[i][1]) ** 2
                    + (testInstance[1] - dataSet[i][2]) ** 2
                    + (testInstance[2] - dataSet[i][3]) ** 2
                    + (testInstance[3] - dataSet[i][4]) ** 2
                ),
                i,
            ]
        )

    distList.sort()

    classes = []
    for i in range(K):

        classes.append(data[distList[i][1]][5])

    result = max(classes, key=classes.count)

    return result


# ____________Naive Bayes Algorithm____________


def Naive_Bayes_Algorithm(dataSet, testInstance):

    def GausDistribution(featureSamples, testInstance):

        sum = 0
        count = 0
        for i in featureSamples:

            sum = sum + i
            count = count + 1

        mean = sum / count

        varianceNumerator = 0
        for i in featureSamples:

            varianceNumerator = varianceNumerator + (i - mean) ** 2

        variance = varianceNumerator / (count - 1)
        standardDeviation = round(math.sqrt(variance), 1)
        GaussianDistribution = (1 / (math.sqrt(2 * math.pi) * standardDeviation)) * (
            math.e ** ((-((testInstance - mean) ** 2)) / (2 * variance))
        )

        return round(GaussianDistribution, 4)

    def Naive_Bayes_Classifier():

        global result

        predictors = []
        responseClass = []
        for j in range(len(dataSet[0])):

            temp = []
            for i in range(1, len(dataSet)):

                if type(dataSet[i][j]) != str:

                    continue

                temp.append(dataSet[i][j])

            if j == len(dataSet[0]) - 1:

                responseClass.append(list(set(temp)))

            else:

                predictors.append(
                    list(set(temp))
                )  # => We get the predictors just for clarity.

        responseClassCount = []
        for j in range(len(responseClass[0])):

            count = 0

            for i in range(len(dataSet)):

                if responseClass[0][j] == dataSet[i][-1]:

                    count += 1

            responseClassCount.append(count)

        # ____________Conditional Probability Of The Features With Categorical Values____________

        responseClassIfCount = []
        temp = []
        for k in range(len(responseClass[0])):

            temp = []
            for i in range(len(testInstance)):

                ifCount = 0

                if type(testInstance[i]) != str:

                    continue

                for j in range(1, len(dataSet)):

                    if (
                        testInstance[i] == dataSet[j][i]
                        and dataSet[j][-1] == responseClass[0][k]
                    ):

                        ifCount += 1

                temp.append(ifCount)

            responseClassIfCount.append(temp)

        # ____________Conditional Probability Of The Features With Continuous Values____________

        continuousValuesPosteriors = []
        temp = []
        for k in range(len(responseClass[0])):

            temp1 = []

            for i in range(len(testInstance)):

                if type(testInstance[i]) == str:

                    continue

                temp = []
                for j in range(1, len(dataSet)):

                    if dataSet[j][-1] == responseClass[0][k]:

                        temp.append(dataSet[j][i])

                Gausdistribution_ForEach_ClassFeature = GausDistribution(
                    temp, testInstance[i]
                )
                temp1.append(Gausdistribution_ForEach_ClassFeature)

            continuousValuesPosteriors.append(temp1)

        allPosteriors = []
        temp = []
        for i in range(len(responseClassCount)):

            temp = []
            for j in range(len(responseClassIfCount[i])):

                temp.append(
                    round((responseClassIfCount[i][j] / responseClassCount[i]), 4)
                )

            allPosteriors.append(temp)

        for i in range(len(responseClass[0])):

            allPosteriors[i] = allPosteriors[i] + continuousValuesPosteriors[i]

        posteriorProbabilityResult = []
        for i in range(len(allPosteriors)):

            productTemp = 1

            for j in range(len(allPosteriors[i])):

                productTemp *= allPosteriors[i][j]

            productTemp *= responseClassCount[i] / (len(dataSet) - 1)
            posteriorProbabilityResult.append(productTemp)

        result = responseClass[0][
            posteriorProbabilityResult.index(max(posteriorProbabilityResult))
        ]

    Naive_Bayes_Classifier()

    return result


# ____________Import Data From Excel____________


def Import_Data_From_Excel(Excel_Name):

    global data, r

    dataframe = openpyxl.load_workbook(Excel_Name)

    dataframe1 = dataframe.active

    data = []
    for row in range(1, dataframe1.max_row):

        temp = []
        for col in dataframe1.iter_cols(1, dataframe1.max_column):

            temp.append(col[row].value)

        data.append(temp)

    r = []
    for row in range(1, dataframe1.max_row):

        temp = []
        for col in dataframe1.iter_cols(6, 6):

            temp.append(col[row].value)

        r.append(temp[0])

    print("\n@@@@@ The Excel Sheet Data @@@@@")

    print("\nThe number of cases:", len(r))
    print("The number of patients with Anemia:", r.count("Anemia"))
    print("The number of healthy people:", r.count("No Anemia"), "\n")


# ____________Use One Algorithm From The Algorithms____________


def Use_Algorithm(AlgName):

    Anemia_Rsult_Number = 0
    Healthy_People_Rsult_Number = 0

    True_Result = 0
    False_Result = 0

    for i in range(len(data)):

        if AlgName == "Naive_Bayes":

            result = Naive_Bayes_Algorithm(data, data[i][:5])

        if AlgName == "KNN":

            result = KNN(data, 5, data[i][1:5])

        if result == "Anemia":

            Anemia_Rsult_Number += 1

        if result == "No Anemia":

            Healthy_People_Rsult_Number += 1

        if r[i] == result:

            True_Result += 1

        if r[i] != result:

            False_Result += 1

    print("\n@@@@@ Program Data With {A} Algorithm @@@@@".format(A=AlgName))

    print("\nAnemia_Rsult_Number:", Anemia_Rsult_Number)
    print("Healthy_People_Rsult_Number:", Healthy_People_Rsult_Number)

    print("\nTrue Rsult nummber:", True_Result)
    print("False Rsult nummber:", False_Result)

    print("\nAccuracy percentage:", "%.2f" % ((True_Result / 1421) * 100), "%\n")


Import_Data_From_Excel("anemia.xlsx")

Use_Algorithm("Naive_Bayes")
Use_Algorithm("KNN")


print("\n@@@@@ One Watching Example @@@@@")
print("\nOne Case example:", data[1])
One_Watched_Test = [15.9, 25.4, 28.3, 72]
KNN(data, 3, One_Watched_Test)
print("\nThe nearest classes are:", classes)
print("\nThe result for One_Watched_Test:", result, "\n")
